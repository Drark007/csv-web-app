# Versão final do aplicativo com todas as modificações solicitadas integradas

import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import logging
from datetime import datetime
from dataclasses import dataclass
from typing import Optional, List
import csv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ==================== CONSTANTES ====================
CSV_EXTENSIONS = ('.csv',)
DEFAULT_ENCODINGS = ('utf-8', 'utf-8-sig', 'latin1', 'iso-8859-1')
LOG_FORMAT = '%(asctime)s - %(levelname)s - %(message)s'
DATE_FORMAT = '%Y_%m_%d'

# ==================== CONFIGURAÇÕES ====================
@dataclass
class Config:
    substituir_virgulas: bool = True
    converter_volume: bool = True
    remover_duplicatas: bool = False
    tentar_encoding_multiplo: bool = True

# ==================== FUNÇÕES AUXILIARES ====================
def selecionar_arquivos_csv() -> Optional[List[str]]:
    root = tk.Tk()
    root.withdraw()

    def on_close():
        if messagebox.askyesno("Confirmar", "Deseja realmente sair?"):
            root.destroy()
            os._exit(0)
    root.protocol("WM_DELETE_WINDOW", on_close)

    messagebox.showinfo("Selecione arquivos", "Escolha os arquivos CSV que deseja processar")
    arquivos = filedialog.askopenfilenames(filetypes=[("Arquivos CSV", "*.csv")])
    root.destroy()

    return list(arquivos) if arquivos else None

def selecionar_arquivo_saida(pasta_origem: Optional[str] = None) -> Optional[str]:
    root = tk.Tk()
    root.withdraw()
    nome_base = datetime.now().strftime(DATE_FORMAT)
    if pasta_origem:
        nome_base += f"_{os.path.basename(pasta_origem)}"
    caminho = filedialog.asksaveasfilename(
        title="Salvar resultado como",
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile=f"{nome_base}.xlsx"
    )
    root.destroy()
    return caminho

def detectar_separador(caminho: str, encoding: str) -> str:
    try:
        with open(caminho, encoding=encoding) as f:
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(f.readline())
            return dialect.delimiter
    except:
        return ';'

def ler_csv_com_encoding(caminho: str) -> Optional[pd.DataFrame]:
    for encoding in DEFAULT_ENCODINGS:
        try:
            sep = detectar_separador(caminho, encoding)
            return pd.read_csv(caminho, sep=sep, encoding=encoding)
        except Exception:
            continue
    return None

def obter_colunas_comuns(arquivos: List[str]) -> Optional[List[str]]:
    colunas_por_arquivo = []
    for arq in arquivos:
        df = ler_csv_com_encoding(arq)
        if df is None:
            continue
        colunas_por_arquivo.append(set(df.columns))

    if not colunas_por_arquivo:
        return None

    intersecao = set.intersection(*colunas_por_arquivo)
    todas_iguais = all(c == colunas_por_arquivo[0] for c in colunas_por_arquivo)
    return list(intersecao) if todas_iguais else None

def selecionar_colunas(colunas: List[str]) -> List[str]:
    root = tk.Tk()
    root.title("Selecionar colunas")
    selecionadas = []
    vars_dict = {}

    tk.Label(root, text="Selecione as colunas que deseja extrair:").pack(pady=10)

    for col in colunas:
        var = tk.BooleanVar(value=True)
        cb = tk.Checkbutton(root, text=col, variable=var)
        cb.pack(anchor='w')
        vars_dict[col] = var

    def confirmar():
        for col, var in vars_dict.items():
            if var.get():
                selecionadas.append(col)
        root.destroy()

    tk.Button(root, text="Confirmar", command=confirmar).pack(pady=10)
    root.mainloop()
    return selecionadas

def processar_csv(caminho: str, config: Config, colunas_desejadas: Optional[List[str]]) -> Optional[pd.DataFrame]:
    df = ler_csv_com_encoding(caminho)
    if df is None or df.empty:
        return None

    if colunas_desejadas:
        df = df[colunas_desejadas]

    if 'volume' in df.columns and config.converter_volume:
        df['volume'] = pd.to_numeric(df['volume'], errors='coerce') / 1000

    if config.substituir_virgulas:
        for col in df.select_dtypes(include=['object']):
            try:
                numeric_test = pd.to_numeric(df[col].str.replace(',', '.'), errors='coerce')
                if numeric_test.notna().sum() > 0:
                    df[col] = numeric_test.map(lambda x: f"{x:.2f}".replace('.', ',') if pd.notna(x) else '')
            except:
                continue

    if config.remover_duplicatas:
        df = df.drop_duplicates()

    return df

def main():
    logging.basicConfig(level=logging.INFO, format=LOG_FORMAT)
    config = Config()

    if not (arquivos := selecionar_arquivos_csv()):
        return

    colunas_comuns = obter_colunas_comuns(arquivos)
    colunas_desejadas = None
    if colunas_comuns:
        messagebox.showinfo("Arquivos compatíveis", "Todos os arquivos possuem as mesmas colunas. Selecione as desejadas.")
        colunas_desejadas = selecionar_colunas(colunas_comuns)

    janela = tk.Tk()
    janela.title("Processando arquivos")

    def ao_fechar():
        if messagebox.askyesno("Confirmar", "Deseja realmente sair?"):
            janela.destroy()
            os._exit(0)
    janela.protocol("WM_DELETE_WINDOW", ao_fechar)

    progresso = tk.DoubleVar()
    barra = ttk.Progressbar(janela, maximum=len(arquivos), variable=progresso, length=400)
    barra.pack(padx=20, pady=10)
    status = tk.Label(janela, text="Iniciando...")
    status.pack()
    janela.update()

    dados = []
    for i, caminho in enumerate(arquivos, 1):
        status.config(text=f"Processando: {os.path.basename(caminho)} ({i}/{len(arquivos)})")
        janela.update()
        df = processar_csv(caminho, config, colunas_desejadas)
        if df is not None:
            dados.append((f"Documento {i}", df))
        progresso.set(i)
        janela.update()

    status.config(text="Finalizado.")
    janela.after(1000, janela.destroy)
    janela.mainloop()

    if not dados:
        messagebox.showerror("Erro", "Nenhum dado processado com sucesso")
        return

    if not (saida := selecionar_arquivo_saida(os.path.dirname(arquivos[0]))):
        return

    with pd.ExcelWriter(saida, engine='openpyxl') as writer:
        linha_atual = 1
        for nome_doc, df in dados:
            df.to_excel(writer, startrow=linha_atual+1, index=False, header=True)
            linha_atual += len(df) + 3

        resumo = pd.DataFrame({
            "Total Registros": [sum(len(df) for _, df in dados)],
            "Arquivos Processados": [len(dados)],
            "Colunas Geradas": [", ".join(dados[0][1].columns)]
        })
        resumo.to_excel(writer, index=False, sheet_name="Resumo")

    # Ajustes no Excel para mesclar títulos dos Documentos
    wb = load_workbook(saida)
    ws = wb.active
    linha_atual = 1
    for nome_doc, df in dados:
        col_range = get_column_letter(1) + str(linha_atual) + ":" + get_column_letter(df.shape[1]) + str(linha_atual)
        ws.merge_cells(col_range)
        ws.cell(row=linha_atual, column=1).value = nome_doc
        ws.cell(row=linha_atual, column=1).alignment = Alignment(horizontal='center')
        linha_atual += len(df) + 3
    wb.save(saida)

    # Mensagem final personalizada
    aviso = tk.Tk()
    aviso.title("Processamento concluído")
    tk.Label(aviso, text=f"Arquivo salvo em:\n{saida}", wraplength=500, justify='left').pack(padx=20, pady=20)
    tk.Button(aviso, text="Fechar", command=aviso.destroy).pack(pady=10)
    aviso.mainloop()

if __name__ == "__main__":
    main()
